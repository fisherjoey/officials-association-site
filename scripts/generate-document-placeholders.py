#!/usr/bin/env python3
"""
Generate labelled placeholder documents for public/documents/.

PLAT-22: the real operational documents that shipped in this repo (fee
schedule, invoice policy, services agreement, scheduling templates) must
be removed. This script regenerates single-page/single-sheet placeholders
at the same paths and extensions so download links and UI keep working,
without pulling in any new runtime dependency (uses only the stdlib PDF
writer below, plus openpyxl which is already installed for the xlsx
side).

Usage:
    python3 scripts/generate-document-placeholders.py

Regenerate anytime the placeholder wording needs to change; the script is
idempotent (it always overwrites the target files).
"""

from __future__ import annotations

import os
import zlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS_DIR = os.path.join(REPO_ROOT, "public", "documents")


# ---------------------------------------------------------------------------
# Minimal hand-written single-page PDF writer (no external dependency).
# ---------------------------------------------------------------------------


def _pdf_escape(text: str) -> str:
    return text.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")


def _wrap(text: str, max_chars: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) > max_chars and current:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines


def make_placeholder_pdf(path: str, title: str, purpose: str, body_lines: list[str]) -> None:
    """Write a single-page US Letter PDF with a title and wrapped body text."""

    content_lines: list[str] = []
    content_lines.append("BT")
    content_lines.append("/F2 20 Tf")
    content_lines.append("72 760 Td")
    content_lines.append(f"({_pdf_escape('SAMPLE DOCUMENT')}) Tj")
    content_lines.append("ET")

    content_lines.append("BT")
    content_lines.append("/F1 13 Tf")
    content_lines.append("72 725 Td")
    content_lines.append(f"({_pdf_escape(title)}) Tj")
    content_lines.append("ET")

    y = 690
    content_lines.append("BT")
    content_lines.append("/F1 11 Tf")
    content_lines.append(f"72 {y} Td")
    content_lines.append("14 TL")
    wrapped_purpose = _wrap(
        "Replace this placeholder with your organisation's own " + purpose + ".",
        90,
    )
    first = True
    for line in wrapped_purpose:
        if first:
            content_lines.append(f"({_pdf_escape(line)}) Tj")
            first = False
        else:
            content_lines.append(f"T* ({_pdf_escape(line)}) Tj")
    content_lines.append("ET")
    y -= 14 * len(wrapped_purpose) + 20

    content_lines.append("BT")
    content_lines.append("/F1 11 Tf")
    content_lines.append(f"72 {y} Td")
    content_lines.append("16 TL")
    first = True
    for line in body_lines:
        wrapped = _wrap(line, 90) or [""]
        for w in wrapped:
            if first:
                content_lines.append(f"({_pdf_escape(w)}) Tj")
                first = False
            else:
                content_lines.append(f"T* ({_pdf_escape(w)}) Tj")
    content_lines.append("ET")

    content_stream = "\n".join(content_lines).encode("latin-1", "replace")
    compressed = zlib.compress(content_stream)

    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
    objects.append(
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R /F2 5 0 R >> >> /Contents 6 0 R >>"
    )
    objects.append(
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica "
        b"/Encoding /WinAnsiEncoding >>"
    )
    objects.append(
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold "
        b"/Encoding /WinAnsiEncoding >>"
    )
    stream_obj = (
        b"<< /Length "
        + str(len(compressed)).encode()
        + b" /Filter /FlateDecode >>\nstream\n"
        + compressed
        + b"\nendstream"
    )
    objects.append(stream_obj)

    out = bytearray()
    out += b"%PDF-1.4\n"
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n".encode()
        out += obj
        out += b"\nendobj\n"

    xref_offset = len(out)
    out += f"xref\n0 {len(objects) + 1}\n".encode()
    out += b"0000000000 65535 f \n"
    for off in offsets[1:]:
        out += f"{off:010d} 00000 n \n".encode()
    out += b"trailer\n"
    out += f"<< /Size {len(objects) + 1} /Root 1 0 R >>\n".encode()
    out += b"startxref\n"
    out += f"{xref_offset}\n".encode()
    out += b"%%EOF"

    with open(path, "wb") as f:
        f.write(out)


# ---------------------------------------------------------------------------
# Minimal single-sheet xlsx writer (openpyxl, already installed).
# ---------------------------------------------------------------------------


def make_placeholder_xlsx(path: str, title: str, purpose: str, columns: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Placeholder"

    ws.merge_cells("A1:E1")
    ws["A1"] = "SAMPLE DOCUMENT — " + title
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A2:E2")
    ws["A2"] = "Replace this placeholder with your organisation's own " + purpose + "."
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:E3")
    ws["A3"] = "This sheet contains no real schedule data."
    ws["A3"].font = Font(italic=True)

    header_row = 5
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.font = Font(bold=True)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    wb.save(path)


def main() -> None:
    make_placeholder_pdf(
        os.path.join(DOCS_DIR, "Fee-Schedule.pdf"),
        title="Fee Schedule",
        purpose="fee schedule",
        body_lines=[
            "This document would normally list per-game or per-event rates paid "
            "to officials, broken down by level, age group, or role.",
            "No real pricing information is included in this sample file.",
        ],
    )

    make_placeholder_pdf(
        os.path.join(DOCS_DIR, "Invoice-Policy.pdf"),
        title="Invoice Policy",
        purpose="invoice policy",
        body_lines=[
            "This document would normally describe how officials submit invoices "
            "for payment, including deadlines, required information, and payment "
            "turnaround times.",
            "No real policy terms are included in this sample file.",
        ],
    )

    make_placeholder_pdf(
        os.path.join(DOCS_DIR, "Officiating-Services-Agreement.pdf"),
        title="Officiating Services Agreement",
        purpose="officiating services agreement",
        body_lines=[
            "This document would normally set out the contractual terms between "
            "an officials association and a league or event organiser for the "
            "provision of officiating services.",
            "No real contract terms are included in this sample file.",
        ],
    )

    make_placeholder_xlsx(
        os.path.join(DOCS_DIR, "League-Scheduling-Template.xlsx"),
        title="League Scheduling Template",
        purpose="league scheduling template",
        columns=["Date", "Time", "Home Team", "Away Team", "Assigned Official(s)"],
    )

    make_placeholder_xlsx(
        os.path.join(DOCS_DIR, "League-Scheduling-Template-Google.xlsx"),
        title="League Scheduling Template (Google Sheets version)",
        purpose="league scheduling template",
        columns=["Date", "Time", "Home Team", "Away Team", "Assigned Official(s)"],
    )

    make_placeholder_xlsx(
        os.path.join(DOCS_DIR, "Tournament-Scheduling-Template.xlsx"),
        title="Tournament Scheduling Template",
        purpose="tournament scheduling template",
        columns=["Date", "Time", "Court/Venue", "Teams", "Assigned Official(s)"],
    )

    make_placeholder_xlsx(
        os.path.join(DOCS_DIR, "Tournament-Scheduling-Template-Google.xlsx"),
        title="Tournament Scheduling Template (Google Sheets version)",
        purpose="tournament scheduling template",
        columns=["Date", "Time", "Court/Venue", "Teams", "Assigned Official(s)"],
    )

    print("Wrote placeholders to", DOCS_DIR)


if __name__ == "__main__":
    main()
